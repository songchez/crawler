from openpyxl import Workbook
from openpyxl import load_workbook


# #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
# load_wb = load_workbook("/Users/tama4/Desktop/gogog/goon.xlsx", data_only=True)
# #시트 이름으로 불러오기
# load_ws = load_wb['Sheet1']
# #셀 주소로 값 출력
# print(load_ws['A1'].value)
# #셀 좌표로 값 출력
# print(load_ws.cell(1,2).value)

write_wb = Workbook()
write_ws = write_wb.active

#행 단위로 추가
write_ws.append([1,2,3])
write_ws.append([1,2,3])
 
#셀 단위로 추가
write_ws.cell(5,5,'5행5열')
write_wb.save("/Users/tama4/Desktop/gogog/goon.xlsx")