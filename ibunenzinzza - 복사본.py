import requests
import json
from openpyxl import Workbook
from openpyxl import load_workbook
import random     
import os

#custom_header을 통해 위장하기
custom_header = {
    'referer' : 'https://m.youtube.com/',
    'user-agent' : 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'  }

#해당 접속 사이트가 아닌 원본데이터가 오는 url 추적. network에서 가지고 온다.
url = "https://m.youtube.com/results?search_query=파스타먹방"

# 1. 밑에 숫자부분을 따오는 코드를 작성해야한다. nope!!
# 2. json 데이터를 가공하는 작업을 해야한다. good!!!

write_wb = Workbook() #활성화
write_ws = write_wb.create_sheet('간다1')
write_ws.append(['상호','대표자','주소','전화번호','등록업'])

def parse_json(num) :
    data = {'checkNum' : num} #json 매개변수(파라미터)전송
    html_1 = requests.get(url, headers=custom_header, data = data)  #custom_header를 사용하지 않으면 접근 불가
    parsdata = html_1.json() #json으로 변환
    jsongo = dict(parsdata[0]) #딕셔너리로 변환
    items=jsongo['strItem'] 
    write_ws.append([jsongo['strSangho'],jsongo['strCeo'],jsongo['strAddr'],jsongo['strTel'],items[:13]]) #작성
    write_wb.save("/Users/tama4/Desktop/gogog/goon.xlsx")

count = 0 
go = 0
for i in range(0,15000):
    alpa=str(random.randint(0,200000))
    parse_json(alpa.zfill(6))
    i += 1 
    go += 1
    print("현재",i,"개째 하고있습니다.")
    if go > 700: #700개마다 저장
        count += 1
        old_name = "/Users/tama4/Desktop/gogog/goon.xlsx"
        new_name = "/Users/tama4/Desktop/gogog/"+"goona" + str(count) + ".xlsx"
        os.rename(old_name, new_name)
        go = 0

print("끝!!!!")

#끝내고 싶을땐 컨드롤+c
